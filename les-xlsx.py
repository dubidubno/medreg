import openpyxl
from pathlib import Path
import sys
import mysql.connector
from mysql.connector import errorcode
import yaml
import io

try:
    configfile = io.open("config.yaml",'r',encoding='utf8')
except:
    sys.exit("failed to open configfile")
try:
    config = yaml.safe_load(configfile)
except:
    sys.exit("failed to read or parse configfile")


data = []
xlsx_file = Path('.', 'Bok3.xlsx');
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active
print (sheet["A1"].value)
print(sheet.max_row, sheet.max_column)

col_names = []
for column in sheet.iter_cols(1, sheet.max_column):
    col_names.append(column[0].value)
   
    
print(col_names)

if col_names == ['Fanenr', 'Navn', 'Mobil', 'Epost', 'Adresse']:
    print ("OK")
else:
    sys.exit("ERROR: Feil i kolonnenavn")
    
print(col_names[0])

for i, row in enumerate(sheet.iter_rows(values_only=True)):
    if i == 0:
        # do nothing
        # for x, field in enumerate(row):
            # print("{}: {}".format(x, field))
            # data[field] = []
        print("hei")
    else:
        # for x, field in enumerate(row):
            # data[col_names[x]].append(field)
            # # print("{}, {}: {}".format(i, col_names[x], field))
        data.append({})
        for x, field in enumerate(row):
            data[i-1][col_names[x]] = field
            # print("{}, {}: {}".format(i, col_names[x], field))
print("read finished")

try:
    mydb = mysql.connector.connect(
        host=config['database-host'],
        user=config['database-user'],
        password=config['database-password'],
        database=config['database-name'],
        # ssl_disabled = True,
    )
except mysql.connector.Error as err:
    if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
        print("Something is wrong with your user name or password")
    elif err.errno == errorcode.ER_BAD_DB_ERROR:
        print("Database does not exist")
    else:
        print(err)
    sys.exit
else:
    print("DB connection ok")

mycursor = mydb.cursor()
insert_stmt = "INSERT INTO medlemmer (fanenr, navn, mobil, epost, adresse) VALUES (%s, %s, %s, %s, %s)"
update_stmt = "UPDATE medlemmer SET updated_time = now(), navn = %s, mobil = %s, epost = %s, adresse = %s WHERE fanenr = %s"
for  i, row in enumerate(data):
    print ("{}:{}".format(i, row['Navn']), end = '')
    mycursor.execute('SELECT * FROM medlemmer WHERE fanenr = %s', (row['Fanenr'],))
    rec = mycursor.fetchone()
    if not rec: 
        # medlem ikke i databasen, bruk insert
        print (" (insert)")
        try:
            mycursor.execute(insert_stmt, (row['Fanenr'], row['Navn'], row['Mobil'], row['Epost'], row['Adresse']))
            mydb.commit()           
        except Exception as e:
            print(e)
            sys.exit("Insert feilet.")
    else:
        # medlem allerede i databasen, bruk update
        print (" (update)")
        mycursor.execute(update_stmt, (row['Navn'], row['Mobil'], row['Epost'], row['Adresse'],row['Fanenr']))
        mydb.commit()

print("The end.")

    


