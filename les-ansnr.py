import openpyxl
from pathlib import Path
import sys
import mysql.connector
from mysql.connector import errorcode
import yaml
import io

configfile_name = "config.yaml"
try:
    configfile = io.open(configfile_name,'r',encoding='utf8')
except:
    sys.exit("Failed to open configfile. ({})".format(configfile_name))
try:
    config = yaml.safe_load(configfile)
except:
    sys.exit("failed to read or parse configfile")


data = []
xlsx_file = Path('.', '2022-01-14 - medlemsliste med ansattnr.xlsx');
try:
    wb_obj = openpyxl.load_workbook(xlsx_file)
except:
    sys.exit("Failed to opeen excel-file.")
sheet = wb_obj.active
print (sheet["A1"].value)
print(sheet.max_row, sheet.max_column)

col_names = []
for column in sheet.iter_cols(1, sheet.max_column):
    col_names.append(column[0].value)
    
print(col_names)

for i, row in enumerate(sheet.iter_rows(values_only=True)):
    if i == 0:
        # do nothing, skipping first row with column names
        print("hei")
    else:
        data.append({})
        data[i-1]['fanenr'] = row[1]
        data[i-1]['ansattnr'] = row[0]
        print("{}: {} {}".format(i, row[0], row[1]))
print("read finished")

try:
    mydb = mysql.connector.connect(
        host=config['database-host'],
        user=config['database-user'],
        password=config['database-password'],
        database=config['database-name'],
        # ssl_disabled = True,
    )
except mysql.connector.Error as err:
    if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
        print("Something is wrong with your user name or password")
    elif err.errno == errorcode.ER_BAD_DB_ERROR:
        print("Database does not exist")
    else:
        print(err)
    sys.exit
else:
    print("DB connection ok")
mycursor = mydb.cursor()

for rec in data:
    if rec['ansattnr'] is not None: 
        mycursor.execute("UPDATE medlemmer SET ansattnr = %s WHERE fanenr = %s",(rec['ansattnr'], rec['fanenr']))
        mydb.commit()

print("The end.")
